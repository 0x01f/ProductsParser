from __future__ import annotations

import os
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .types import Product


DEFAULT_HEADERS_RU = [
    "Название",
    "Цена",
    "Ссылка",
    "Ссылка на изображение",
    "Описание",
]


def _ensure_sheet(wb_path: Optional[str]) -> tuple[Workbook, Worksheet, bool]:
    """
    Returns (workbook, sheet, is_new_file)
    """
    if wb_path and os.path.exists(wb_path):
        wb = load_workbook(wb_path)
        ws = wb.active
        return wb, ws, False
    wb = Workbook()
    ws = wb.active
    return wb, ws, True


def write_products_to_excel(
    products: Iterable[Product],
    out_path: str,
    template_path: Optional[str] = None,
    headers: Optional[list[str]] = None,
) -> None:
    headers = headers or DEFAULT_HEADERS_RU

    # If template is provided and exists, start from it
    wb_path_to_open = template_path if (template_path and os.path.exists(template_path)) else None
    wb, ws, is_new = _ensure_sheet(wb_path_to_open)

    # If we started from template, we will save to out_path (not overwrite template)
    # Ensure headers exist if sheet seems empty
    if ws.max_row == 1 and ws.max_column == 1 and (ws.cell(row=1, column=1).value is None):
        for col_idx, title in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).value = title

    start_row = ws.max_row + 1 if ws.max_row >= 1 else 1
    for idx, p in enumerate(products, start=start_row):
        ws.cell(row=idx, column=1).value = p.name
        ws.cell(row=idx, column=2).value = p.price
        ws.cell(row=idx, column=3).value = p.url
        ws.cell(row=idx, column=4).value = p.image_url
        ws.cell(row=idx, column=5).value = p.description

    # Always save to out_path
    wb.save(out_path)

